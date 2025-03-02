from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bot.utils.data_processing.date_converter import format_date

class ExcelCRUD:
    current_dir = Path(__file__).resolve().parent.parent.parent
    file_path = current_dir / "files"

    @staticmethod
    async def create_new_excel(table_id: int, table_name: str, tg_id: int):
        try:
            file_path = ExcelCRUD.file_path / f"{table_name}_{table_id}_{tg_id}.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = f'Клиенты "{table_name}"'

            headers = ['ID', 'Username / TG_ID', 'Название OS', 'Дата окончания', 'Цена']
            for i, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=i, value=header)
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            sheet.column_dimensions["A"].width = 15
            sheet.column_dimensions["B"].width = 25
            sheet.column_dimensions["C"].width = 25
            sheet.column_dimensions["D"].width = 20
            sheet.column_dimensions["E"].width = 15

            workbook.save(file_path)
            return file_path

        except Exception as e:
            print(f"Error creating new excel file: {e}")
            return None

    @staticmethod
    async def add_client_to_existing_excel(data, table_name: str, table_id: int, tg_id: int):
        file_path = ExcelCRUD.file_path / f"{table_name}_{table_id}_{tg_id}.xlsx"

        if file_path.exists():
            workbook = load_workbook(file_path)
            sheet = workbook.active
            first_empty_row = sheet.max_row + 1

            row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") if first_empty_row % 2 == 0 else PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            for i, word in enumerate(data, start=1):
                cell = sheet.cell(row=first_empty_row, column=i, value=word)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = row_fill

            workbook.save(file_path)
            return True
        return False

    @staticmethod
    async def delete_excel_file(table_id: int, table_name: str, tg_id: int):
        try:
            file_path = ExcelCRUD.file_path / f"{table_name}_{table_id}_{tg_id}.xlsx"

            if file_path.exists():
                file_path.unlink()
                return True
            else:
                return False
        except Exception as e:
            return False

    @classmethod
    async def get_excel_file(cls, table_id: int, table_name: str, tg_id: int, clients, connections):
        file_path = cls.file_path / f"{table_name}_{table_id}_{tg_id}.xlsx"

        if not file_path.exists():
            await cls.create_new_excel(table_id=table_id, table_name=table_name, tg_id=tg_id)

        row_num = 2
        for client in clients:
            client_connections = [conn for conn in connections if conn.client_id == client.id]
            for connection in client_connections:
                client_data = [
                    client.id,
                    client.username,
                    connection.os_name,
                    format_date(connection.date_to),
                    connection.price
                ]
                await cls.add_client_to_existing_excel(
                    data=client_data, table_name=table_name, table_id=table_id, tg_id=tg_id
                )
                row_num += 1

        return file_path
