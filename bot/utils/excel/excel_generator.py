from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from bot.utils.data_processing.date_converter import format_date


class ExcelCRUD:
    current_dir = Path(__file__).resolve().parent.parent.parent
    file_path = current_dir / "files"

    @staticmethod
    async def create_new_excel(table_id: int, table_name: str, tg_id: int):
        try:
            file_path = ExcelCRUD.file_path / f"{table_name}_{table_id}_{tg_id}.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = table_name

            header_items = ["TG", "ЦЕНА", "ДАТА НАЧАЛА", "ДАТА КОНЦА", "ЗАДЕРЖКА"]
            for i, header in enumerate(header_items, start=1):
                cell = sheet.cell(row=1, column=i, value=header)
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True)

            sheet.column_dimensions["A"].width = 15
            sheet.column_dimensions["B"].width = 10
            sheet.column_dimensions["C"].width = 25
            sheet.column_dimensions["D"].width = 25
            sheet.column_dimensions["E"].width = 15

            workbook.save(file_path)
            return True

        except:
            return False

    @staticmethod
    async def add_client_to_existing_excel(
        data, table_name: str, table_id: int, tg_id: int
    ):
        file_path = ExcelCRUD.file_path / f"{table_name}_{table_id}_{tg_id}.xlsx"

        if file_path.exists():
            workbook = load_workbook(file_path)
            sheet = workbook.active
            first_empty_row = sheet.max_row + 1

            for i, word in enumerate(data, start=1):
                cell = sheet.cell(row=first_empty_row, column=i, value=word)
                cell.alignment = Alignment(horizontal="center")

            workbook.save(file_path)

            return True
        return False

    @staticmethod
    async def delete_excel_file(table_id: int, table_name: str, tg_id: int):
        try:
            file_path = ExcelCRUD.file_path / f"{table_name}_{table_id}_{tg_id}.xlsx"

            if file_path.exists():
                file_path.unlink()
                return True
            else:
                return False
        except Exception as e:
            return False

    @classmethod
    async def get_excel_file(cls, table_id: int, table_name: str, tg_id: int, clients):
        file_path = cls.file_path / f"{table_name}_{table_id}_{tg_id}.xlsx"

        await cls.create_new_excel(
            table_id=table_id, table_name=table_name, tg_id=tg_id
        )

        for client in clients:
            client_data = [client.username, client.price, format_date(client.date_from), format_date(client.date_to), client.days_late if client.days_late != 0 else '']
            await cls.add_client_to_existing_excel(
                data=client_data, table_name=table_name, table_id=table_id, tg_id=tg_id
            )

        return file_path
