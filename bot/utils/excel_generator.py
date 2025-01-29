from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


class ExcelCRUD:
    current_dir = Path(__file__).resolve().parent.parent.parent
    file_path = current_dir / "bot" / "files"

    @classmethod
    async def create_new_excel(cls, table_id: int, table_name: str, tg_id: int):
        try:
            file_path = cls.file_path / f"{table_name}_{table_id}_{tg_id}.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = table_name

            header_items = ["TG", "ЦЕНА", "СРОК"]
            for i, header in enumerate(header_items, start=1):
                cell = sheet.cell(row=1, column=i, value=header)
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True)

            sheet.column_dimensions["A"].width = 15
            sheet.column_dimensions["B"].width = 10
            sheet.column_dimensions["C"].width = 25

            workbook.save(file_path)
            return True
        
        except:
            return False

    @classmethod
    async def add_words_to_existing_excel(cls, data, table_name: str, table_id: int, tg_id: int):
        file_path = cls.file_path / f"{table_name}_{table_id}_{tg_id}.xlsx"

        if file_path.exists():
            workbook = load_workbook(file_path)
            sheet = workbook.active
            first_empty_row = sheet.max_row + 1

            for i, word in enumerate(data, start=1):
                cell = sheet.cell(row=first_empty_row, column=i, value=word)
                cell.alignment = Alignment(horizontal="center")

            workbook.save(file_path)

            return True
        return False